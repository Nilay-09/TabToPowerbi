import re
import pandas as pd
from openpyxl import load_workbook

def write_dataframes_to_excel(sheet_data, excel_file):
    """
    Writes a dictionary of DataFrames to an Excel workbook.
    Applies date formatting ('yyyy-mm-dd') for datetime columns.
    """
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        sheet_names = []
        for sheet_name, df in sheet_data.items():
            # Sanitize the sheet name (max 31 chars, remove invalid characters)
            safe_sheet_name = re.sub(r'[\/:*?"<>|]', '_', sheet_name)[:31]
            base_name = safe_sheet_name
            counter = 1
            while safe_sheet_name in sheet_names:
                suffix = f"_{counter}"
                safe_sheet_name = base_name[:31 - len(suffix)] + suffix
                counter += 1
            sheet_names.append(safe_sheet_name)
            df.to_excel(writer, sheet_name=safe_sheet_name, index=False)
            print(f"✅ Written sheet '{safe_sheet_name}' with {len(df)} rows to Excel.")
    # Adjust Excel cell formatting for date columns using openpyxl
    wb = load_workbook(excel_file)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header = [cell.value for cell in ws[1]]
        for col_idx, col_name in enumerate(header, start=1):
            if ws.max_row > 1:
                cell_value = ws.cell(row=2, column=col_idx).value
                if isinstance(cell_value, (pd.Timestamp, str)):
                    try:
                        pd.to_datetime(cell_value, errors='raise')
                        for row in range(2, ws.max_row + 1):
                            ws.cell(row=row, column=col_idx).number_format = 'yyyy-mm-dd'
                    except Exception:
                        pass
    wb.save(excel_file)
    print(f"✅ Data written and formatted in Excel file: {excel_file}")
    return sheet_names
